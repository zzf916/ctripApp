# -*- coding: utf-8 -*-
# File : config.py
# Author: Off
# Date : 2022/2/12
# Desc : 配置文件

import configparser
import os
import re
import time
from itertools import cycle

import openpyxl

from comm.logs import Logging

log = Logging().logger

ini_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '../config'))
timestamp = time.strftime('%Y-%m-%d %H:%M:%S')


class MyParser(configparser.ConfigParser):
    """Inherit from built-in class: ConfigParser"""

    def optionxform(self, optionstr):
        """Rewrite without lower()"""

        return optionstr


def is_dict(value):
    try:
        if isinstance(eval(value), dict):
            return eval(value)
        else:
            return value
    except Exception:
        return value


# 获取配置文件
def getConfigData(section):
    cf = MyParser()
    cf.read(ini_path + '/config.ini', encoding='utf-8')
    dict_ini = {}
    for i in cf.items(section):
        key = i[0]
        value = is_dict(i[1])
        if value == 'Ture':
            value = True
        dict_ini[key] = value
    return dict_ini


# Excel文件读写数据
class Excel(object):
    def __init__(self, *, workbook_path, sheet):
        self.workbook_path = workbook_path
        self.workbook = openpyxl.load_workbook(workbook_path)
        self.sheet = self.workbook[sheet]
        self.r_max = self.sheet.max_row
        self.c_max = self.sheet.max_column

    def ExcelR(self):
        """
        一次读取整个文件
        :return:list [(酒店名，酒店编号)]
        """
        list_ = []
        for r in range(1, self.r_max + 1):
            hotel_data = (self.sheet.cell(row=r, column=3).value, self.sheet.cell(row=r, column=2).value)
            list_.append(hotel_data)
        return list_

    def numR(self, num: int):
        """
        :param num:读取的数据数
        :return: list [(酒店名，酒店编号, 行数)]
        """
        list_ = []
        for r in range(1, self.r_max + 1):
            if len(list_) < num and self.sheet.cell(row=r, column=5).value is None:
                hotel_data = (self.sheet.cell(row=r, column=3).value, self.sheet.cell(row=r, column=2).value, r)
                list_.append(hotel_data)

        return list_

    def CycleR(self):
        """
        一次读一行
        :return:
        """
        for r in range(1, self.r_max + 1):
            hotel_data = (self.sheet.cell(row=r, column=3).value, self.sheet.cell(row=r, column=2).value, r)

            yield hotel_data

    def ExcelW(self, args, row, column=5):
        """
        Excel写数据
        :param column:
        :param args: 写入的数据
        :param row: 行数
        :return:
        """
        try:
            self.sheet.cell(row=row, column=column).value = args
            self.workbook.save(self.workbook_path)
            self.workbook.close()
        except PermissionError as e:
            log.info("写入Excel失败:{}".format(e))


if __name__ == '__main__':
    # print(getConfigData('trip'))
    workbook = getConfigData('Excel').get('workbook')
    # print(workbook)
    # data = Excel(workbook_path=workbook, sheet='Sheet1').ExcelR()
    # print(data)
    # row = data.get('row')
    Excel(workbook_path=workbook, sheet='Sheet1').ExcelW(timestamp, 1)
    # i = Excel(workbook_path=workbook, sheet='Sheet1').CycleR()
    # n = cycle(data)

    # while True:
    #     print(next(n))
    #     print('\n')
    data = Excel(workbook_path=workbook, sheet='Sheet1').numR(2)
    # print(len(data))
    print(data)
